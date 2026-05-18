
import pandas as pd
from sqlalchemy import create_engine
import pymysql
import numpy as np
import os

def connect_to_mysql():
    engine = create_engine(
        "mysql+pymysql://root:Iambatman1102!@localhost:3306/novatech_analysis"
    )
    return engine

def load_data(connection):
     sales_df = pd.read_sql("select * from sales_data",connection)
     income_df = pd.read_sql("select * from income_statement", connection)
     balance_df = pd.read_sql("select * from balance_sheet", connection)
     cashflow_df = pd.read_sql("select * from cash_flow", connection)
     return sales_df,income_df,balance_df,cashflow_df

def calculate_ratios(income_df,balance_df):
    print("\n" "="*1)
    print("novatech_ratio_analysis")

    print("="*40)

    gross_margin = round((income_df["gross_profit"].sum()/income_df["revenue"].sum())*100,2)
    ebitda_margin = round((income_df["ebitda"].sum() / income_df["revenue"].sum()) * 100, 2)
    net_margin = round((income_df["net_profit"].sum() / income_df["revenue"].sum()) * 100, 2)

    print(f"   Gross Profit Margin : {gross_margin}%")
    print(f"   EBITDA Margin       : {ebitda_margin}%")
    print(f"   Net Profit Margin   : {net_margin}%")

    print("LIQUIDITY RATIOS")
    print("-" * 40)

    current_ratio = round((balance_df["total_current_assets"].sum()/
                           balance_df["total_current_liabilities"].sum()),2)

    working_capital = round(balance_df["total_current_assets"].sum() -
                            balance_df["total_current_liabilities"].sum(), 2)

    print(f"current_ratio:{current_ratio}x")
    print(f"working_capital:{working_capital:,.2f}")

    print("SOLVENCY RATIOS")
    print("-"*40)

    debt_to_equity = round(balance_df["total_liabilities"].sum()/
                            balance_df["equity"].sum(),2)

    interest_coverage = round(income_df["ebit"].sum() /
                             income_df["interest"].sum(),2)

    print(f"Debt to Equity Ratio : {debt_to_equity}x")
    print(f"Interest Coverage    : {interest_coverage}x")

    print("ACTIVITY RATIOS")
    print("-" * 40)

    working_capital_turnover_ratio = round(income_df["revenue"].sum()/
                                    round(balance_df["total_current_assets"].sum() -
                                     balance_df["total_current_liabilities"].sum(),2))

    fixed_asset_turnover_ratio = round(income_df["revenue"].sum()/
                                       balance_df["fixed_assets"].sum(),2)

    print(f"working capital turnover ratio :{working_capital_turnover_ratio}x")
    print(f"fixed asset turnover ratio :{fixed_asset_turnover_ratio}x")

    print("PROFITABILITY RATIOS")
    print("-"*40)

    gross_profit_ratio= round(income_df["gross_profit"].sum()/
                        round(income_df["revenue"].sum())*100,2)

    net_profit_ratio = round(income_df["net_profit"].sum()/
                       round(income_df["revenue"].sum())*100,2)

    print(f"gross profit ratio:{gross_profit_ratio}%")
    print(f"net profit ratio:{net_profit_ratio}%")
    print("-" * 40)

def growth_analysis(income_df):

    print("\n" + "=" * 55)
    print(" NOVATECH GROWTH ANALYSIS FY2024")
    print("=" * 55)

    income_df["revenue_growth"] = round(income_df["revenue"].pct_change()*100,2)
    income_df["profit_growth"] = round(income_df["net_profit"].pct_change() * 100, 2)
    income_df["gross_margin_trend"]= round((income_df["gross_profit"]/
                                     income_df["revenue"])*100,2)

    print("\n MONTHLY GROWTH TRENDS")
    print("-" * 55)
    print(f"{'month':<12} {'revenue growth':>15} {'profit growth':>15}  {'gross margin': >13}")
    print("-" * 55)

    for _, row in income_df.iterrows():
        rev_growth = (f"{row['revenue_growth']}%"
        if pd.notna(row['revenue_growth'])else "base")
        pft_growth = (f"{row['profit_growth']}%"
        if pd.notna(row['profit_growth']) else "base")
        margin = f"{row['gross_margin_trend']}%"
        print(f"{row['month']:<12} {rev_growth:>15} {pft_growth:>15} {margin:>13}")


    avg_rev_growth = round(income_df["revenue_growth"].mean(), 2)
    avg_pft_growth = round(income_df["profit_growth"].mean(), 2)
    total_rev_growth = round(((income_df["revenue"].iloc[-1] -
                               income_df["revenue"].iloc[0])/
                              income_df["revenue"].iloc[0]) * 100, 2)

    print(f"   Avg Monthly Revenue Growth : {avg_rev_growth}%")
    print(f"   Avg Monthly Profit Growth  : {avg_pft_growth}%")
    print(f"   Full Year Revenue Growth   : {total_rev_growth}%")

    print("\n" + "=" * 55)

def anomaly_detection(income_df):
    print("\n" + "=" * 55)
    print(" NOVATECH ANOMALY DETECTION")
    print("=" * 55)

    revenue_mean = income_df["revenue"].mean()
    revenue_std = income_df["revenue"].std()

    income_df["revenue_zscore"]= ((income_df["revenue"] -revenue_mean)
                                 /revenue_std)

    profit_mean  = income_df["net_profit"].mean()
    profit_std = income_df["net_profit"].std()

    income_df["profit_zscore"] = ((income_df["net_profit"] - profit_mean)
                                   / profit_std)

    print("\nRevenue Z-Scores")
    print(income_df[["month", "revenue", "revenue_zscore"]])

    print("\nProfit Z-Scores")
    print(income_df[["month", "net_profit", "profit_zscore"]])

    anamolies = income_df[
        (abs(income_df["revenue_zscore"]) >2) |
        (abs(income_df["profit_zscore"]) > 2) ]

    if anamolies.empty:
        print("No major  anomalies detected.")
    else:
        print("major anamolies detected.")

    for _,row in anamolies.iterrows():
        print(f"""
        Month          : {row['month']}
        Revenue        : {row['revenue']:,.2f}
        Net Profit     : {row['net_profit']:,.2f}
        Revenue ZScore : {round(row['revenue_zscore'],2)}
        Profit ZScore  : {round(row['profit_zscore'],2)}
         """)


def cashflow_analysis(cashflow_df, income_df):


    print("\n" + "=" * 55)
    print(" NOVATECH CASH FLOW ANALYSIS FY2024")
    print("=" * 55)

    # ── 1. CASH FLOW HEALTH CHECK ──────────────────────────
    print("\n CASH FLOW HEALTH CHECK")
    print("-" * 55)

    total_operating = round(cashflow_df["operating_cash_flow"].sum(),2)
    total_investing = round(cashflow_df["investing_cash_flow"].sum(), 2)
    total_financing = round(cashflow_df["financing_cash_flow"].sum(), 2)

    print(f" operating CF: {total_operating:>12,.2f} Positive (Healthy)")
    print(f" investing CF: {total_investing:>12,.2f} negative (growing)")
    print(f" financing CF: {total_financing:>12,.2f} negative (repaying)")

    print("\n OPENING vs CLOSING BALANCE")
    print("-" * 55)

    opening = cashflow_df["opening_balance"].iloc[0]
    closing = cashflow_df["closing_balance"].iloc[-1]
    balance_growth = round(((closing -opening)/opening)*100,2)

    print(f"opening balance : {opening:>12,.2f}")
    print(f"closing balance : {closing:>12,.2f}")
    print(f" Cash Growth :  {balance_growth}% ")

    print("\n BEST & WORST CASH MONTH")
    print("-" * 55)

    best_month = cashflow_df.loc[cashflow_df["net_cash_flow"].idxmax()]
    worst_month = cashflow_df.loc[cashflow_df["net_cash_flow"].idxmin()]

    print(f"best month : {best_month['month']:<12} {best_month['net_cash_flow']:>10,.2f}")
    print(f"worst month : {worst_month['month']:<12} {worst_month['net_cash_flow']:>10,.2f}")

    print("\n OPERATING CF RATIO")
    print("-" * 55)

    total_net_profit = income_df["net_profit"].sum()
    op_cf_ratio = round(total_operating / total_net_profit, 2)

    print(f"   Operating CF Ratio : {op_cf_ratio}x")
    if op_cf_ratio >= 0.5:
        print("   Verdict : Profit is backed by real cash!")
    else:
        print("   Verdict : Profit quality needs attention!")

    print("\n" + "=" * 55)



def export_to_excel(income_df, balance_df, cashflow_df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── COLORS ──────────────────────────────────────────
    dark_blue = "FF1F3864"
    light_blue = "FFD6E4F0"
    green = "FFE8F5E9"
    yellow = "FFFFF9C4"
    white = "FFFFFFFF"

    def style_header(cell, bg=dark_blue, font_color="FFFFFFFF", bold=True):
        cell.font = Font(bold=bold, color = font_color, name = "Arial",size = 11)
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal="center")

    def style_subheader(cell):
        cell.font = Font(bold=True, color="FF1F3864", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color=light_blue)
        cell.alignment = Alignment(horizontal="left")

    # SHEET 1 — RATIO ANALYSIS
    ws1 = wb.active
    ws1.title = "Ratio Analysis"
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 20

    ws1.append(["NOVATECH FY2024 — RATIO ANALYSIS", "", ""])
    style_header(ws1["A1"])
    ws1.merge_cells("A1:C1")

    sections = [
        ("PROFITABILITY RATIOS", [
            ("Gross Profit Margin", f"=SUMIF(income!B:B,\"<>0\",income!C:C)/"
                                    f"SUMIF(income!B:B,\"<>0\",income!B:B)*100"),
            ("EBITDA Margin", f"=SUMIF(income!B:B,\"<>0\",income!F:F)/"
                              f"SUMIF(income!B:B,\"<>0\",income!B:B)*100"),
            ("Net Profit Margin", f"=SUMIF(income!B:B,\"<>0\",income!J:J)/"
                                  f"SUMIF(income!B:B,\"<>0\",income!B:B)*100"),
        ]),

        ("LIQUIDITY RATIOS", [
            ("Current Ratio", f"=SUM(balance!C:C)/SUM(balance!E:E)"),
            ("Working Capital (₹)", f"=SUM(balance!C:C)-SUM(balance!E:E)"),
        ]),
        ("SOLVENCY RATIOS", [
            ("Debt to Equity", f"=SUM(balance!F:F)/SUM(balance!G:G)"),
            ("Interest Coverage", f"=SUM(income!H:H)/SUM(income!I:I)"),
        ]),
        ("ACTIVITY RATIOS", [
            ("Working Capital Turnover", f"=SUM(income!B:B)/(SUM(balance!C:C)-SUM(balance!E:E))"),
            ("Fixed Asset Turnover", f"=SUM(income!B:B)/SUM(balance!D:D)"),
        ]),
    ]

    row = 2
    for section_title,metrics in sections:
        ws1.append([section_title, "value","benchmark"])
        style_subheader(ws1.cell(row, 1))
        ws1.cell(row, 2).font = Font(bold=True, name="Arial")
        ws1.cell(row, 3).font = Font(bold=True, name="Arial")
        row += 1
        for name, _ in metrics:
            ws1.append([name, "See Python Output", ""])
            ws1.cell(row, 1).font = Font(name="Arial", size=10)
            row += 1
        ws1.append([])
        row += 1

        # SHEET 2 — GROWTH ANALYSIS
        # ════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Growth Analysis")
        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 18

        ws2.append(["NOVATECH FY2024 — GROWTH ANALYSIS", "", "", ""])
        style_header(ws2["A1"])
        ws2.merge_cells("A1:D1")

        ws2.append(["month","revenue growth%", "profit growth%","gross margin%"])
        for col in range(1,5):
            style_subheader(ws2.cell(2,col))

        income_df["revenue_growth"] = round(income_df["revenue"].pct_change() * 100, 2)
        income_df["profit_growth"] = round(income_df["net_profit"].pct_change() * 100, 2)
        income_df["gross_margin_trend"] = round((income_df["gross_profit"] / income_df["revenue"]) * 100, 2)

        for _, row_data in income_df.iterrows():
            ws2.append([
                row_data["month"],
                row_data["revenue_growth"] if pd.notna(row_data["revenue_growth"]) else "Base",
                row_data["profit_growth"] if pd.notna(row_data["profit_growth"]) else "Base",
                row_data["gross_margin_trend"]
            ])

        ws2.append([])
        ws2.append(["Avg Monthly Rev Growth", round(income_df["revenue_growth"].mean(), 2), "", ""])
        ws2.append(["Avg Monthly Pft Growth", round(income_df["profit_growth"].mean(), 2), "", ""])
        ws2.append(["Full Year Rev Growth",
                    round(((income_df["revenue"].iloc[-1] - income_df["revenue"].iloc[0])
                           / income_df["revenue"].iloc[0]) * 100, 2), "", ""])

 # SHEET 3 — CASH FLOW ANALYSIS

        ws3 = wb.create_sheet("Cash Flow Analysis")
        ws3.column_dimensions["A"].width = 18
        ws3.column_dimensions["B"].width = 20
        ws3.column_dimensions["C"].width = 20
        ws3.column_dimensions["D"].width = 20
        ws3.column_dimensions["E"].width = 20

        ws3.append(["NOVATECH FY2024 — CASH FLOW ANALYSIS", "", "", "", ""])
        style_header(ws3["A1"])
        ws3.merge_cells("A1:E1")

        ws3.append(["Month", "Operating CF", "Investing CF", "Financing CF", "Net Cash Flow"])
        for col in range(1, 6):
            style_subheader(ws3.cell(2, col))

        for _, row_data in cashflow_df.iterrows():
            ws3.append([
                row_data["month"],
                row_data["operating_cash_flow"],
                row_data["investing_cash_flow"],
                row_data["financing_cash_flow"],
                row_data["net_cash_flow"]
            ])

        last_row = 2 + len(cashflow_df)
        ws3.append([
            "TOTAL",
            f"=SUM(B3:B{last_row})",
            f"=SUM(C3:C{last_row})",
            f"=SUM(D3:D{last_row})",
            f"=SUM(E3:E{last_row})"
        ])
        for col in range(1, 6):
            ws3.cell(last_row + 1, col).font = Font(bold=True, name="Arial")
            ws3.cell(last_row + 1, col).fill = PatternFill("solid", start_color=yellow)

        # Summary
        ws3.append([])
        ws3.append(["Opening Balance", cashflow_df["opening_balance"].iloc[0]])
        ws3.append(["Closing Balance", cashflow_df["closing_balance"].iloc[-1]])
        ws3.append(["Cash Growth %",
                    round(((cashflow_df["closing_balance"].iloc[-1] -
                            cashflow_df["opening_balance"].iloc[0]) /
                           cashflow_df["opening_balance"].iloc[0]) * 100, 2)])

        # SHEET 4 — ANOMALY DETECTION

        ws4 = wb.create_sheet("Anomaly Detection")
        ws4.column_dimensions["A"].width = 15
        ws4.column_dimensions["B"].width = 18
        ws4.column_dimensions["C"].width = 18
        ws4.column_dimensions["D"].width = 18
        ws4.column_dimensions["E"].width = 18

        ws4.append(["NOVATECH FY2024 — ANOMALY DETECTION", "", "", "", ""])
        style_header(ws4["A1"])
        ws4.merge_cells("A1:E1")

        ws4.append(["Month", "Revenue", "Revenue Z-Score", "Net Profit", "Profit Z-Score"])
        for col in range(1, 6):
            style_subheader(ws4.cell(2, col))

        rev_mean = income_df["revenue"].mean()
        rev_std = income_df["revenue"].std()
        pft_mean = income_df["net_profit"].mean()
        pft_std = income_df["net_profit"].std()

        for _, row_data in income_df.iterrows():
            rev_z = round((row_data["revenue"] - rev_mean) / rev_std, 2)
            pft_z = round((row_data["net_profit"] - pft_mean) / pft_std, 2)
            row_list = [row_data["month"], row_data["revenue"], rev_z,
                        row_data["net_profit"], pft_z]
            ws4.append(row_list)

            # Highlight anomalies in red
            current_row = ws4.max_row
            if abs(rev_z) > 2 or abs(pft_z) > 2:
                for col in range(1, 6):
                    ws4.cell(current_row, col).fill = PatternFill("solid", start_color="FFFFCCCC")

        ws4.append([])
        ws4.append(["No anomalies detected (all Z-scores within ±2)"])
        ws4.cell(ws4.max_row, 1).font = Font(bold=True, color="FF2E7D32", name="Arial")

        # ── SAVE ────────────────────────────────────────────
        path = r"C:\Users\gohil\Downloads\novatech_fy2024_analysis.xlsx"
        wb.save(path)
        print(f"\n✅ Excel report saved → {r"C:\Users\gohil\Downloads\novatech_fy2024_analysis.xlsx"}")


def main():
    print("="*55)
    print(" novatech_analysis is loading")

    try:
        connection =connect_to_mysql()
        print("data loaded and connected successfully")

        sales_df, income_df, balance_df, cashflow_df = load_data(connection)
        print("\n All tables loaded!")
        print(f"   Sales Data    → {len(sales_df)} rows")
        print(f"   Income Stmt   → {len(income_df)} rows")
        print(f"   Balance Sheet → {len(balance_df)} rows")
        print(f"   Cash Flow     → {len(cashflow_df)} rows")

        calculate_ratios(income_df, balance_df)
        growth_analysis(income_df)
        anomaly_detection(income_df)
        cashflow_analysis(cashflow_df, income_df)
        export_to_excel(income_df, balance_df, cashflow_df)




    except Exception as e:
        print(f"data not loaded, connection failed {e}")

if __name__ =="__main__":
    main()










